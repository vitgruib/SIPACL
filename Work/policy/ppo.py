# docs and experiment results can be found at https://docs.cleanrl.dev/rl-algorithms/ppo/#ppo_continuous_actionpy
import os
import random
import sys

# Ensure Work is on path so "custom" and relative paths resolve when run from repo root or Work
_work_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _work_dir not in sys.path:
    sys.path.insert(0, _work_dir)
import time
import traceback
from datetime import datetime
import scenic
from dataclasses import dataclass
from typing import Optional
from custom.custom_simulator import CustomMetaDriveSimulation, CustomMetaDriveSimulator 


import gymnasium as gym
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import tyro
from torch.distributions.normal import Normal
from custom.custom_gym import MetaDriveEnv


class _NoOpWriter:
    """No-op stand-in for SummaryWriter when TensorBoard logging is disabled."""

    def add_text(self, *args, **kwargs): pass
    def add_scalar(self, *args, **kwargs): pass
    def close(self): pass


try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = Workbook = None  # optional: pip install openpyxl


@dataclass
class Args:
    exp_name: str = os.path.basename(__file__)[: -len(".py")]
    """the name of this experiment"""
    seed: int = 1
    """seed of the experiment"""
    torch_deterministic: bool = True
    """if toggled, `torch.backends.cudnn.deterministic=False`"""
    cuda: bool = True
    """if toggled, cuda will be enabled by default"""
    track: bool = False
    """if toggled, this experiment will be tracked with Weights and Biases"""
    wandb_project_name: str = "cleanRL"
    """the wandb's project name"""
    wandb_entity: str = None
    """the entity (team) of wandb's project"""
    capture_video: bool = False
    """whether to capture videos of the agent performances (check out `videos` folder)"""
    upload_model: bool = False
    """whether to upload the saved model to huggingface"""
    hf_entity: str = ""
    """the user or org name of the model repository from the Hugging Face Hub"""


    # Scenic specific arguments (rl_mobil scenario + Town06; controllers in scenarios/controllers/)
    scenic_file: str = "./scenarios/rl_mobil.scenic"
    """path to the Scenic program defining the environment"""
    max_steps: int = 1000
    """the maximum number of steps for any given episode; episodes truncate after this"""
    model: str = "scenic.simulators.metadrive.model"
    """underlying model for the scenic file"""
    map: str = "./CARLA/Town06.net.xml"
    """Sumo/OpenDRIVE map for the env (rl_mobil uses Town06)"""
    sampler_type: str = "random"
    """ sampling type for generating scenes"""
    save_model: bool = True
    """whether to save model to the `runs/{run_name}` folder"""
    evaluate_model: bool = False
    """if True, skip training and only evaluate the model at model_to_evaluate_path"""
    eval_after_train: bool = True
    """if True, run evaluation for eval_episodes after training (same run)"""
    eval_episodes: int = 10
    """number of episodes to run when evaluating (evaluate_model or eval_after_train)"""
    eval_max_steps: int = 1000
    """fixed max steps per eval episode so returns are comparable; -1 uses training max_steps"""
    model_to_evaluate_path: str = "runs/ACL_MetaDrive__ppo__1__resample-1__1772936486.pt"
    """Path for model to evaluate when evaluate_model is True"""
    render: Optional[bool] = None
    """override rendering (default: true for eval, false otherwise)"""
    render3d: bool = False
    """override 3D rendering (default: true)"""

    # Prioritized level replay (MetaDriveEnv)0
    replay_resample_prob: float = .5
    """probability of resampling from buffer vs new scene; use -1 to disable replay"""
    buffer_dir: str = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "buffer")
    """directory for scene buffer (scene_*.bin and buffer_*.npy); default is Work/buffer"""
    resume_from_buffer: bool = False
    """load buffer state from buffer_dir on init (continue a previous run)"""

    # Algorithm specific arguments
    env_id: str = "ACL_MetaDrive"
    """the id of the environment"""
    total_timesteps: int = 500000
    """total timesteps of the experiments"""
    learning_rate: float = 3e-4
    """the learning rate of the optimizer"""
    num_envs: int = 1
    """the number of parallel game environments"""
    num_steps: int = 4096
    """the number of steps to run in each environment per policy rollout"""
    anneal_lr: bool = True
    """Toggle learning rate annealing for policy and value networks"""
    gamma: float = 0.99
    """the discount factor gamma"""
    gae_lambda: float = 0.95
    """the lambda for the general advantage estimation"""
    num_minibatches: int = 32
    """the number of mini-batches"""
    update_epochs: int = 10
    """the K epochs to update the policy"""
    norm_adv: bool = True
    """Toggles advantages normalization"""
    clip_coef: float = 0.2
    """the surrogate clipping coefficient"""
    clip_vloss: bool = True
    """Toggles whether or not to use a clipped loss for the value function, as per the paper."""
    ent_coef: float = 0.01
    """coefficient of the entropy"""
    vf_coef: float = 0.5
    """coefficient of the value function"""
    max_grad_norm: float = 0.5
    """the maximum norm for the gradient clipping"""
    target_kl: float = None
    """the target KL divergence threshold"""

    # to be filled in runtime
    batch_size: int = 0
    """the batch size (computed in runtime)"""
    minibatch_size: int = 0
    """the mini-batch size (computed in runtime)"""
    num_iterations: int = 0
    """the number of iterations (computed in runtime)"""


def make_env(env_id, idx, capture_video, run_name, gamma, max_steps_override=None):
    def thunk():
        steps = max_steps_override if max_steps_override is not None else args.max_steps
        try:
            scenario = (scenic.scenarioFromFile(args.scenic_file,
                                            model=args.model,
                                            mode2D=True,
                                            params={"verifaiSamplerType": args.sampler_type}))
        except AssertionError as e:
            print("AssertionError loading Scenic scenario:", file=sys.stderr)
            print("  args.scenic_file =", repr(args.scenic_file), "  cwd =", repr(os.getcwd()), file=sys.stderr)
            print("  exception args:", e.args, file=sys.stderr)
            traceback.print_exc()
            raise
        #shape was [100 200   3]
        # OBS needs to be updated
        render_flag = args.render if args.render is not None else args.evaluate_model
        render3d_flag = args.render3d
        env = MetaDriveEnv(
            scenario=scenario,
            simulator=CustomMetaDriveSimulator(
                sumo_map=args.map,
                max_steps=steps,
                render=render_flag,
                render3D=render3d_flag,
                timestep=0.1,
            ),
            max_steps=steps,
            replay_resample_prob=args.replay_resample_prob,
            buffer_dir=args.buffer_dir,
            resume_from_buffer=args.resume_from_buffer,
        )

        # Keep flattening because policy network expects a flat Box observation tensor.
        env = gym.wrappers.FlattenObservation(env)
        env = gym.wrappers.RecordEpisodeStatistics(env)
        return env

    return thunk


def layer_init(layer, std=np.sqrt(2), bias_const=0.0):
    torch.nn.init.orthogonal_(layer.weight, std)
    torch.nn.init.constant_(layer.bias, bias_const)
    return layer




class Agent(nn.Module):
    def __init__(self, obs_dim, action_dim):
        super().__init__()
        self.critic = nn.Sequential(
            layer_init(nn.Linear(obs_dim, 64)),
            nn.Tanh(),
            layer_init(nn.Linear(64, 64)),
            nn.Tanh(),
            layer_init(nn.Linear(64, 1), std=1.0),
        )
        self.actor_mean = nn.Sequential(
            layer_init(nn.Linear(obs_dim, 64)),
            nn.Tanh(),
            layer_init(nn.Linear(64, 64)),
            nn.Tanh(),
            layer_init(nn.Linear(64, action_dim), std=0.01),
        )
        # Start with lower exploration noise to reduce control jitter.
        self.actor_logstd = nn.Parameter(torch.full((1, action_dim), -1.0))

    def get_value(self, x):
        return self.critic(x)

    def get_action_and_value(self, x, action=None):
        action_mean = self.actor_mean(x)
        action_logstd = self.actor_logstd.expand_as(action_mean)
        action_std = torch.exp(action_logstd)
        probs = Normal(action_mean, action_std)
        if action is None:
            pre_tanh = probs.rsample()
            action = torch.tanh(pre_tanh)
        else:
            # action is already squashed to (-1, 1); invert tanh for log-prob.
            eps = 1e-6
            clipped = torch.clamp(action, -1.0 + eps, 1.0 - eps)
            pre_tanh = 0.5 * torch.log((1 + clipped) / (1 - clipped))

        log_prob = probs.log_prob(pre_tanh)
        # Tanh correction
        log_prob = log_prob - torch.log(1 - action * action + 1e-6)
        log_prob = log_prob.sum(1)
        return action, log_prob, probs.entropy().sum(1), self.critic(x)


if __name__ == "__main__":
    args = tyro.cli(Args)
    # Resolve paths relative to Work (parent of policy/) so they work regardless of cwd
    if not os.path.isabs(args.scenic_file):
        args.scenic_file = os.path.normpath(os.path.join(_work_dir, args.scenic_file))
    if not os.path.isabs(args.map):
        args.map = os.path.normpath(os.path.join(_work_dir, args.map))
    if args.model_to_evaluate_path and not os.path.isabs(args.model_to_evaluate_path):
        args.model_to_evaluate_path = os.path.normpath(os.path.join(_work_dir, args.model_to_evaluate_path))
    args.batch_size = int(args.num_envs * args.num_steps)
    args.minibatch_size = int(args.batch_size // args.num_minibatches)
    args.num_iterations = max(1, args.total_timesteps // args.batch_size)
    actual_timesteps = args.num_iterations * args.batch_size
    print(f"Training: total_timesteps={args.total_timesteps}, batch_size={args.batch_size}, num_iterations={args.num_iterations} -> actual steps = {actual_timesteps}, replay_resample_prob={args.replay_resample_prob}")
    run_name = f"p{args.replay_resample_prob}s{args.sampler_type}{datetime.now().strftime('%H%M')}"
    results_excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "runs_results.xlsx")
    episode_returns_log, episode_lengths_log = [], []
    if args.track:
        import wandb

        wandb.init(
            project=args.wandb_project_name,
            entity=args.wandb_entity,
            sync_tensorboard=False,
            config=vars(args),
            name=run_name,
            monitor_gym=True,
            save_code=True,
        )
    writer = _NoOpWriter()

    # TRY NOT TO MODIFY: seeding
    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(args.seed)
        torch.cuda.manual_seed_all(args.seed)
    torch.backends.cudnn.deterministic = args.torch_deterministic

    device = torch.device("cuda" if torch.cuda.is_available() and args.cuda else "cpu")
    print(f"Using device: {device}" + (f" ({torch.cuda.get_device_name(0)})" if device.type == "cuda" else ""))


    # env setup
    envs = gym.vector.SyncVectorEnv(
        [make_env(args.env_id, i, args.capture_video, run_name, args.gamma) for i in range(args.num_envs)]
    )
    assert isinstance(envs.single_action_space, gym.spaces.Box), "only continuous action space is supported"

    next_obs_np, _ = envs.reset(seed=args.seed)
    obs_shape = tuple(next_obs_np.shape[1:])
    obs_dim = int(np.prod(obs_shape))
    action_dim = int(np.prod(envs.single_action_space.shape))

    agent = Agent(obs_dim, action_dim).to(device)
    optimizer = optim.Adam(agent.parameters(), lr=args.learning_rate, eps=1e-5)

    if args.evaluate_model:
        agent.load_state_dict(torch.load(args.model_to_evaluate_path, map_location=device))
        agent.eval()
        obs = torch.Tensor(next_obs_np).to(device)
        episode_returns = np.zeros(args.num_envs, dtype=np.float32)
        completed = 0
        while completed < args.eval_episodes:
            with torch.no_grad():
                # Sampled evaluation: use stochastic policy action (tanh-bounded).
                action, _, _, _ = agent.get_action_and_value(obs)
            next_obs, reward, terminations, truncations, _ = envs.step(action.cpu().numpy())
            episode_returns += reward
            done = np.logical_or(terminations, truncations)
            for i, d in enumerate(done):
                if d:
                    print(f"eval_return: {episode_returns[i]}")
                    completed += 1
                    episode_returns[i] = 0.0
                    if completed >= args.eval_episodes:
                        break
            obs = torch.Tensor(next_obs).to(device)
        envs.close()
        writer.close()
        raise SystemExit(0)

    # ALGO Logic: Storage setup
    obs = torch.zeros((args.num_steps, args.num_envs) + obs_shape).to(device)
    actions = torch.zeros((args.num_steps, args.num_envs) + envs.single_action_space.shape).to(device)
    logprobs = torch.zeros((args.num_steps, args.num_envs)).to(device)
    rewards = torch.zeros((args.num_steps, args.num_envs)).to(device)
    dones = torch.zeros((args.num_steps, args.num_envs)).to(device)
    values = torch.zeros((args.num_steps, args.num_envs)).to(device)

    # TRY NOT TO MODIFY: start the game
    global_step = 0
    start_time = time.time()
    progress_interval = max(1, args.total_timesteps // 100)
    next_progress = progress_interval
    next_obs = torch.Tensor(next_obs_np).to(device)
    next_done = torch.zeros(args.num_envs).to(device)
    final_sps = 0
    # Manual episode tracking (env infos often missing in vector env); used for Excel/logs.
    episode_returns_buf = np.zeros(args.num_envs, dtype=np.float64)
    episode_lengths_buf = np.zeros(args.num_envs, dtype=np.int32)

    for iteration in range(1, args.num_iterations + 1):
        # Annealing the rate if instructed to do so.
        if args.anneal_lr:
            frac = 1.0 - (iteration - 1.0) / args.num_iterations
            lrnow = frac * args.learning_rate
            optimizer.param_groups[0]["lr"] = lrnow

        for step in range(0, args.num_steps):
            global_step += args.num_envs
            while global_step >= next_progress:
                pct = min(100, int(100 * next_progress / args.total_timesteps))
                print(f"progress: {pct}% ({next_progress}/{args.total_timesteps})")
                next_progress += progress_interval
            obs[step] = next_obs
            dones[step] = next_done

            # ALGO LOGIC: action logic
            with torch.no_grad():
                action, logprob, _, value = agent.get_action_and_value(next_obs)
                values[step] = value.flatten()
            actions[step] = action
            logprobs[step] = logprob

            # TRY NOT TO MODIFY: execute the game and log data.
            next_obs, reward, terminations, truncations, infos = envs.step(action.cpu().numpy())
            next_done = np.logical_or(terminations, truncations)
            rewards[step] = torch.tensor(reward).to(device).view(-1)
            next_obs, next_done = torch.Tensor(next_obs).to(device), torch.Tensor(next_done).to(device)

            # Track episode return/length ourselves (vector env often omits episode in infos).
            episode_returns_buf += reward
            episode_lengths_buf += 1
            for i in range(args.num_envs):
                if next_done[i].item():
                    r_val = float(episode_returns_buf[i])
                    l_val = int(episode_lengths_buf[i])
                    reason = infos.get("termination_reason", "unknown")
                    outcome = infos.get("outcome", "unknown")
                    if hasattr(reason, "__getitem__") and not isinstance(reason, str):
                        reason = reason[i] if i < len(reason) else "unknown"
                    if hasattr(outcome, "__getitem__") and not isinstance(outcome, str):
                        outcome = outcome[i] if i < len(outcome) else "unknown"
                    if outcome == "won":
                        summary = "you won (victim crashed or went off road)"
                    elif outcome == "lost":
                        if reason == "crash":
                            summary = "you lost (ego crashed)"
                        elif reason == "out_of_road":
                            summary = "you lost (ego left the road)"
                        elif reason == "terminated":
                            summary = "you lost (ego crashed, off road, or all victims passed you)"
                        else:
                            summary = "you lost"
                    elif outcome == "truncated":
                        summary = "truncated (hit max steps)"
                    else:
                        summary = f"ended (reason={reason}, outcome={outcome})"
                    episode_returns_log.append(r_val)
                    episode_lengths_log.append(l_val)
                    print(
                        f"episode_end | {summary} | reward={r_val:.4f}, length={l_val}, global_step={global_step}"
                    )
                    writer.add_scalar("charts/episodic_return", r_val, global_step)
                    writer.add_scalar("charts/episodic_length", l_val, global_step)
                    episode_returns_buf[i] = 0.0
                    episode_lengths_buf[i] = 0

        # bootstrap value if not done
        with torch.no_grad():
            next_value = agent.get_value(next_obs).reshape(1, -1)
            advantages = torch.zeros_like(rewards).to(device)
            lastgaelam = 0
            for t in reversed(range(args.num_steps)):
                if t == args.num_steps - 1:
                    nextnonterminal = 1.0 - next_done
                    nextvalues = next_value
                else:
                    nextnonterminal = 1.0 - dones[t + 1]
                    nextvalues = values[t + 1]
                delta = rewards[t] + args.gamma * nextvalues * nextnonterminal - values[t]
                advantages[t] = lastgaelam = delta + args.gamma * args.gae_lambda * nextnonterminal * lastgaelam
            returns = advantages + values

        # flatten the batch
        b_obs = obs.reshape((-1,) + obs_shape)
        b_logprobs = logprobs.reshape(-1)
        b_actions = actions.reshape((-1,) + envs.single_action_space.shape)
        b_advantages = advantages.reshape(-1)
        b_returns = returns.reshape(-1)
        b_values = values.reshape(-1)

        # Optimizing the policy and value network
        b_inds = np.arange(args.batch_size)
        clipfracs = []
        for epoch in range(args.update_epochs):
            np.random.shuffle(b_inds)
            for start in range(0, args.batch_size, args.minibatch_size):
                end = start + args.minibatch_size
                mb_inds = b_inds[start:end]

                _, newlogprob, entropy, newvalue = agent.get_action_and_value(b_obs[mb_inds], b_actions[mb_inds])
                logratio = newlogprob - b_logprobs[mb_inds]
                ratio = logratio.exp()

                with torch.no_grad():
                    # calculate approx_kl http://joschu.net/blog/kl-approx.html
                    old_approx_kl = (-logratio).mean()
                    approx_kl = ((ratio - 1) - logratio).mean()
                    clipfracs += [((ratio - 1.0).abs() > args.clip_coef).float().mean().item()]

                mb_advantages = b_advantages[mb_inds]
                if args.norm_adv:
                    mb_advantages = (mb_advantages - mb_advantages.mean()) / (mb_advantages.std() + 1e-8)

                # Policy loss
                pg_loss1 = -mb_advantages * ratio
                pg_loss2 = -mb_advantages * torch.clamp(ratio, 1 - args.clip_coef, 1 + args.clip_coef)
                pg_loss = torch.max(pg_loss1, pg_loss2).mean()

                # Value loss
                newvalue = newvalue.view(-1)
                if args.clip_vloss:
                    v_loss_unclipped = (newvalue - b_returns[mb_inds]) ** 2
                    v_clipped = b_values[mb_inds] + torch.clamp(
                        newvalue - b_values[mb_inds],
                        -args.clip_coef,
                        args.clip_coef,
                    )
                    v_loss_clipped = (v_clipped - b_returns[mb_inds]) ** 2
                    v_loss_max = torch.max(v_loss_unclipped, v_loss_clipped)
                    v_loss = 0.5 * v_loss_max.mean()
                else:
                    v_loss = 0.5 * ((newvalue - b_returns[mb_inds]) ** 2).mean()

                entropy_loss = entropy.mean()
                loss = pg_loss - args.ent_coef * entropy_loss + v_loss * args.vf_coef

                optimizer.zero_grad()
                loss.backward()
                nn.utils.clip_grad_norm_(agent.parameters(), args.max_grad_norm)
                optimizer.step()

            if args.target_kl is not None and approx_kl > args.target_kl:
                break

        y_pred, y_true = b_values.cpu().numpy(), b_returns.cpu().numpy()
        var_y = np.var(y_true)
        explained_var = np.nan if var_y == 0 else 1 - np.var(y_true - y_pred) / var_y

        # TRY NOT TO MODIFY: record rewards for plotting purposes
        writer.add_scalar("charts/learning_rate", optimizer.param_groups[0]["lr"], global_step)
        writer.add_scalar("losses/value_loss", v_loss.item(), global_step)
        writer.add_scalar("losses/policy_loss", pg_loss.item(), global_step)
        writer.add_scalar("losses/entropy", entropy_loss.item(), global_step)
        writer.add_scalar("losses/old_approx_kl", old_approx_kl.item(), global_step)
        writer.add_scalar("losses/approx_kl", approx_kl.item(), global_step)
        writer.add_scalar("losses/clipfrac", np.mean(clipfracs), global_step)
        writer.add_scalar("losses/explained_variance", explained_var, global_step)
        print("SPS:", int(global_step / (time.time() - start_time)))
        final_sps = int(global_step / (time.time() - start_time))
        writer.add_scalar("charts/SPS", final_sps, global_step)

    eval_returns_log = []
    if args.save_model:
        metadrive_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        runs_dir = os.path.join(metadrive_dir, "runs")
        os.makedirs(runs_dir, exist_ok=True)
        # Save model with same succinct name as run_name: p<prob>s<sampler><HHMM>.pt
        model_path = os.path.join(runs_dir, f"{run_name}.pt")
        torch.save(agent.state_dict(), model_path)
        model_path_abs = os.path.abspath(model_path)
        print(f"Run saved: model at {model_path_abs}")
        # from cleanrl_utils.evals.ppo_eval import evaluate

        # episodic_returns = evaluate(
        #     model_path,
        #     make_env,
        #     args.env_id,
        #     eval_episodes=10,
        #     run_name=f"{run_name}-eval",
        #     Model=Agent,
        #     device=device,
        #     gamma=args.gamma,
        # )
        # for idx, episodic_return in enumerate(episodic_returns):
        #     writer.add_scalar("eval/episodic_return", episodic_return, idx)

        # if args.upload_model:
        #     from cleanrl_utils.huggingface import push_to_hub

        #     repo_name = f"{args.env_id}-{args.exp_name}-seed{args.seed}"
        #     repo_id = f"{args.hf_entity}/{repo_name}" if args.hf_entity else repo_name
        #     push_to_hub(args, episodic_returns, repo_id, "PPO", f"runs/{run_name}", f"videos/{run_name}-eval")

    if args.eval_after_train:
        agent.eval()
        next_obs_np, _ = envs.reset(seed=args.seed + 1)
        obs = torch.Tensor(next_obs_np).to(device)
        episode_returns_eval = np.zeros(args.num_envs, dtype=np.float32)
        steps_this_episode = np.zeros(args.num_envs, dtype=np.int32)
        completed = 0
        # Enforce eval_max_steps in the loop when env may never terminate (e.g. max_steps=-1).
        # Only applied when num_envs==1 so we can reset and get a clean next episode.
        use_eval_step_cap = args.eval_max_steps > 0 and args.num_envs == 1
        while completed < args.eval_episodes:
            with torch.no_grad():
                action, _, _, _ = agent.get_action_and_value(obs)
            next_obs, reward, terminations, truncations, _ = envs.step(action.cpu().numpy())
            episode_returns_eval += reward
            steps_this_episode += 1
            done = np.logical_or(terminations, truncations)
            force_reset_needed = False
            if use_eval_step_cap and steps_this_episode[0] >= args.eval_max_steps and not done[0]:
                done[0] = True  # force end so we record and reset
                force_reset_needed = True
            for i, d in enumerate(done):
                if d:
                    r = float(episode_returns_eval[i])
                    eval_returns_log.append(r)
                    print(f"eval_return: {r}")
                    writer.add_scalar("eval/episodic_return", r, completed)
                    completed += 1
                    episode_returns_eval[i] = 0.0
                    steps_this_episode[i] = 0
                    if completed >= args.eval_episodes:
                        break
            # Only when we force-ended does the env not auto-reset; then start next episode.
            if force_reset_needed and completed < args.eval_episodes:
                next_obs_np, _ = envs.reset(seed=args.seed + 1 + completed)
                obs = torch.Tensor(next_obs_np).to(device)
            else:
                obs = torch.Tensor(next_obs).to(device)
        print(f"eval_after_train: ran {args.eval_episodes} episodes (same envs as training, max_steps={args.max_steps}, eval_step_cap={args.eval_max_steps if use_eval_step_cap else 'off'})")

    if load_workbook is not None and Workbook is not None:
        mean_return = float(np.mean(episode_returns_log)) if episode_returns_log else np.nan
        std_return = float(np.std(episode_returns_log)) if len(episode_returns_log) > 1 else (0.0 if episode_returns_log else np.nan)
        mean_length = float(np.mean(episode_lengths_log)) if episode_lengths_log else np.nan
        mean_eval = float(np.mean(eval_returns_log)) if eval_returns_log else ""
        std_eval = float(np.std(eval_returns_log)) if len(eval_returns_log) > 1 else (0.0 if len(eval_returns_log) == 1 else "")
        header = [
            "run_name", "time", "seed", "replay_resample_prob", "total_timesteps", "num_envs", "learning_rate", "gamma",
            "exp_name", "eval_episodes", "eval_max_steps",
            "mean_episodic_return", "std_episodic_return", "mean_episodic_length", "num_episodes", "SPS",
            "episodic_returns",
            "mean_eval_return", "std_eval_return", "num_eval_episodes", "eval_returns",
        ]
        episodic_returns_str = str(episode_returns_log) if episode_returns_log else ""
        eval_returns_str = str(eval_returns_log) if eval_returns_log else ""
        row = [
            run_name,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            args.seed,
            args.replay_resample_prob,
            args.total_timesteps,
            args.num_envs,
            args.learning_rate,
            args.gamma,
            args.exp_name,
            args.eval_episodes,
            args.eval_max_steps if args.eval_max_steps > 0 else "",
            mean_return,
            std_return,
            mean_length,
            len(episode_returns_log),
            final_sps,
            episodic_returns_str,
            mean_eval,
            std_eval,
            len(eval_returns_log),
            eval_returns_str,
        ]
        try:
            if os.path.isfile(results_excel_path):
                wb = load_workbook(results_excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(header)
            ws.append(row)
            wb.save(results_excel_path)
            print(f"Run saved: results appended to {results_excel_path}")
        except Exception as e:
            print(f"Could not write Excel: {e}")
    else:
        print("Excel logging skipped: pip install openpyxl")

    print("Run saved.")
    envs.close()
    writer.close()
