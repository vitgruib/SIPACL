#!/usr/bin/env python3
"""
Visualize runs from runs_results.xlsx:
  1) Training results: bar chart with confidence intervals
  2) Testing (eval) results: bar chart with confidence intervals
  3) One line per run: episode (x) vs return (y)
Usage: python visualize_runs.py [-n N] [--all] [-f path/to/runs_results.xlsx]
"""
import argparse
import ast
import os
import sys

import numpy as np

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import matplotlib.pyplot as plt
except ImportError:
    print("Install matplotlib: pip install matplotlib", file=sys.stderr)
    sys.exit(1)


# Column indices from ppo.py header
HEADER = [
    "run_name", "time", "seed", "replay_resample_prob", "total_timesteps", "num_envs", "learning_rate", "gamma",
    "exp_name", "eval_episodes", "eval_max_steps",
    "mean_episodic_return", "std_episodic_return", "mean_episodic_length", "num_episodes", "SPS",
    "episodic_returns",
    "mean_eval_return", "std_eval_return", "num_eval_episodes", "eval_returns",
]


def _num(x):
    if x is None or x == "":
        return np.nan
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _parse_episodic_returns(s):
    """Parse episodic_returns from Excel (e.g. '[1.2, 3.4]') to list of floats."""
    if s is None or (isinstance(s, str) and s.strip() == ""):
        return []
    if isinstance(s, (list, tuple)):
        return [float(x) for x in s]
    try:
        out = ast.literal_eval(s)
        return [float(x) for x in out] if out else []
    except (ValueError, SyntaxError, TypeError):
        return []


def load_runs(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    runs = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        d = {}
        for i, h in enumerate(HEADER):
            if i < len(row):
                d[h] = row[i]
            else:
                d[h] = None
        runs.append(d)
    return runs


def select_runs(runs, n, use_all):
    """Return runs to plot: all in row order, or top n by eval/train return."""
    if use_all or n <= 0:
        return runs
    def key(r):
        v = _num(r.get("mean_eval_return"))
        if np.isnan(v):
            v = _num(r.get("mean_episodic_return"))
        return v if not np.isnan(v) else -np.inf
    sorted_runs = sorted(runs, key=key, reverse=True)
    return sorted_runs[:n]


def short_label(run, i):
    """Short label for a run (run_name is long)."""
    run_name = run.get("run_name") or ""
    if "__" in run_name:
        parts = run_name.split("__")
        if len(parts) >= 4:
            return f"{parts[1]}_{parts[2]}_{parts[3]}"
    return run_name[:20] if run_name else f"Run {i}"


def _ci_half_width(std, n_episodes):
    """
    95% confidence interval half-width for the *mean* (not raw spread).
    SEM = std / sqrt(n); 95% CI half-width = 1.96 * SEM.
    If n is missing or < 2, fall back to ±std (no SEM scaling).
    """
    std = np.asarray(std, dtype=float)
    n = np.asarray(n_episodes, dtype=float)
    use_sem = (n >= 2) & ~np.isnan(n)
    half = np.where(use_sem, 1.96 * std / np.sqrt(n), std)
    return np.where(np.isnan(half), 0.0, half)


def _data_ylim(means, err_half_widths, padding_frac=0.15):
    """Y limits that zoom in on the data range (means ± error) with padding."""
    valid = ~np.isnan(means)
    if not np.any(valid):
        return 0.0, 1.0
    err = np.where(np.isnan(err_half_widths) | (err_half_widths < 0), 0.0, err_half_widths)
    lo = (means[valid] - err[valid]).min()
    hi = (means[valid] + err[valid]).max()
    r = hi - lo
    if r <= 0:
        r = max(abs(lo) * 0.1, 1e-9)
    pad = r * padding_frac
    return lo - pad, hi + pad


def main():
    parser = argparse.ArgumentParser(description="Visualize runs from runs_results.xlsx")
    parser.add_argument("-n", type=int, default=10, help="Number of runs to show (default: 10). Ignored if --all.")
    parser.add_argument("--all", action="store_true", help="Plot all runs (no limit)")
    parser.add_argument("-f", "--file", default=None, help="Path to runs_results.xlsx (default: policy/runs_results.xlsx)")
    args = parser.parse_args()

    if args.file:
        path = args.file
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(script_dir, "runs_results.xlsx")

    if not os.path.isfile(path):
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)

    runs = load_runs(path)
    if not runs:
        print("No data rows in Excel.")
        sys.exit(0)

    selected = select_runs(runs, args.n, args.all)
    n = len(selected)

    mean_train = np.array([_num(r.get("mean_episodic_return")) for r in selected])
    std_train = np.array([_num(r.get("std_episodic_return")) for r in selected])
    num_episodes = np.array([_num(r.get("num_episodes")) for r in selected])
    mean_eval = np.array([_num(r.get("mean_eval_return")) for r in selected])
    std_eval = np.array([_num(r.get("std_eval_return")) for r in selected])
    num_eval = np.array([_num(r.get("num_eval_episodes")) for r in selected])
    labels = [short_label(r, i) for i, r in enumerate(selected)]

    # 95% CI for the mean: half-width = 1.96 * (std / sqrt(n)). Uses raw std if n missing.
    err_train = _ci_half_width(std_train, num_episodes)
    err_eval = _ci_half_width(std_eval, num_eval)
    err_train = np.where(np.isnan(err_train), 0.0, err_train)
    err_eval = np.where(np.isnan(err_eval), 0.0, err_eval)

    fig, axes = plt.subplots(3, 1, figsize=(max(10, n * 0.5), 10), sharex=False)

    x = np.arange(n)
    width = 0.6

    # 1) Training results: bar chart with 95% CI of the mean
    ax1 = axes[0]
    mask = ~np.isnan(mean_train)
    if np.any(mask):
        ax1.bar(x[mask], mean_train[mask], width, yerr=err_train[mask], capsize=3, color="C0", alpha=0.8, error_kw={"linewidth": 1})
    ax1.set_ylabel("Return")
    ax1.set_title("Training results (mean ± 95% CI)")
    y1_min, y1_max = _data_ylim(mean_train, err_train)
    ax1.set_ylim(y1_min, y1_max)
    if y1_min <= 0 <= y1_max:
        ax1.axhline(0, color="gray", linewidth=0.5)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha="right")

    # 2) Testing (eval) results: bar chart with 95% CI of the mean
    ax2 = axes[1]
    mask = ~np.isnan(mean_eval)
    if np.any(mask):
        ax2.bar(x[mask], mean_eval[mask], width, yerr=err_eval[mask], capsize=3, color="C1", alpha=0.8, error_kw={"linewidth": 1})
    ax2.set_ylabel("Return")
    ax2.set_title("Testing (eval) results (mean ± 95% CI)")
    y2_min, y2_max = _data_ylim(mean_eval, err_eval)
    ax2.set_ylim(y2_min, y2_max)
    if y2_min <= 0 <= y2_max:
        ax2.axhline(0, color="gray", linewidth=0.5)
    ax2.set_xticks(x)
    ax2.set_xticklabels(labels, rotation=45, ha="right")

    # 3) One line per run: x = episode index, y = return
    ax3 = axes[2]
    all_vals = []
    for i, r in enumerate(selected):
        returns = _parse_episodic_returns(r.get("episodic_returns"))
        if not returns:
            continue
        episodes = np.arange(1, len(returns) + 1, dtype=float)
        ax3.plot(episodes, returns, "-", label=labels[i], alpha=0.8, linewidth=1.5)
        all_vals.extend(returns)
    ax3.set_xlabel("Episode")
    ax3.set_ylabel("Return")
    ax3.set_title("Training return by episode (one line per run)")
    if all_vals:
        arr = np.array(all_vals)
        y3_min, y3_max = np.nanmin(arr), np.nanmax(arr)
        r = y3_max - y3_min
        pad = max(r * 0.1, 1e-9) if r > 0 else 1.0
        ax3.set_ylim(y3_min - pad, y3_max + pad)
        if y3_min <= 0 <= y3_max:
            ax3.axhline(0, color="gray", linewidth=0.5)
    ax3.legend(loc="best", fontsize="small")
    ax3.grid(True, alpha=0.3)

    plt.tight_layout()
    out_path = os.path.join(os.path.dirname(path), "runs_visualization.png")
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    print(f"Saved: {out_path}")
    plt.show()


if __name__ == "__main__":
    main()
